from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


source = Path("/home/ubuntu/upload/MatrizdeTrade-SMP2026.xlsx")
workbook = load_workbook(source, read_only=True, data_only=True)

for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    nonempty = [row for row in rows if any(value not in (None, "") for value in row)]
    header = next((row for row in nonempty[:10] if sum(value not in (None, "") for value in row) >= 2), ())
    test_rows = sum(1 for row in nonempty if any("teste" in str(value).lower() for value in row if value is not None))
    print(f"SHEET|{sheet_name}|rows={len(nonempty)}|test_rows={test_rows}|header={list(header)[:12]}")

for sheet_name in ("REGIONAIS26", "Fornecedores"):
    if sheet_name not in workbook.sheetnames:
        continue
    sheet = workbook[sheet_name]
    rows = [row for row in sheet.iter_rows(values_only=True) if any(value not in (None, "") for value in row)]
    print(f"\nSAMPLE|{sheet_name}")
    for row in rows[:25]:
        print(list(row)[:12])
